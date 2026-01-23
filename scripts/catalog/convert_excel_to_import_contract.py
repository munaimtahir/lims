#!/usr/bin/env python3
"""
Excel Contract Adapter: Converts authoritative Excel to importer contract format.

Reads the authoritative Excel file with sheets:
- Tests, Parameters (mapping), ParameterMaster, ReferenceRanges

Outputs a new Excel file with sheets:
- Tests, Parameters, Mapping, ReferenceRanges

This script ensures the Excel format matches what the importer expects.
"""

import sys
import re
import openpyxl
from openpyxl import Workbook
from pathlib import Path
from decimal import Decimal, InvalidOperation


def safe_decimal(value):
    """Convert value to Decimal, return None if invalid."""
    if value is None or str(value).strip() == '':
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def safe_int(value, default=0):
    """Convert value to int, return default if invalid."""
    if value is None or str(value).strip() == '':
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def safe_str(value, default=''):
    """Convert value to string, return default if None."""
    if value is None:
        return default
    return str(value).strip()


def get_header_map(sheet):
    """Get column index map from first row."""
    headers = {}
    if not sheet or sheet.max_row < 1:
        return headers
    for i, cell in enumerate(sheet[1], 1):
        if cell.value:
            key = str(cell.value).strip().lower().replace(' ', '_').replace('(', '').replace(')', '')
            headers[key] = i
    return headers


def safe_get(row, headers, keys, default=None):
    """Get value from row using header map, trying multiple key variations."""
    for key in keys:
        if key in headers:
            col_idx = headers[key] - 1  # Convert to 0-based
            if 0 <= col_idx < len(row):
                val = row[col_idx]
                if val is not None and str(val).strip() != '':
                    return val
    return default


def convert_excel(input_path, output_path):
    """
    Convert authoritative Excel to importer contract format.
    
    Args:
        input_path: Path to authoritative Excel file
        output_path: Path to output Excel file
    """
    print(f"Reading authoritative Excel: {input_path}")
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    
    # Validate required sheets exist
    required_sheets = ['Tests', 'Parameters', 'ParameterMaster']
    missing = [s for s in required_sheets if s not in wb_in.sheetnames]
    if missing:
        print(f"ERROR: Missing required sheets: {missing}")
        return False
    
    # Create output workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet
    
    # Track data for validation
    test_ids = set()
    parameter_ids = set()
    test_param_mappings = set()
    
    # ===== 1. PROCESS TESTS SHEET =====
    print("Processing Tests sheet...")
    tests_sheet = wb_in['Tests']
    tests_headers = get_header_map(tests_sheet)
    
    ws_tests = wb_out.create_sheet("Tests")
    ws_tests.append([
        'test_id', 'test_code', 'legacy_test_code', 'test_name', 
        'category', 'sample_type', 'price', 'turnaround_time'
    ])
    
    for row_num, row in enumerate(tests_sheet.iter_rows(min_row=2, values_only=True), 2):
        test_id = safe_int(safe_get(row, tests_headers, ['test_id', 'id']))
        if not test_id:
            continue
        
        test_code = safe_str(safe_get(row, tests_headers, ['test_code', 'code']))
        legacy_test_code = safe_str(safe_get(row, tests_headers, ['legacy_test_code', 'legacy_code']))
        test_name = safe_str(safe_get(row, tests_headers, ['test_name', 'name']))
        category = safe_str(safe_get(row, tests_headers, ['category', 'department']), 'General')
        sample_type = safe_str(safe_get(row, tests_headers, ['sample_type', 'specimen']), 'Serum')
        price = safe_decimal(safe_get(row, tests_headers, ['price', 'cost'])) or Decimal('0')
        turnaround_time = safe_int(safe_get(row, tests_headers, ['turnaround_time', 'tat_hours', 'tat']), 24)
        
        # Infer sample_type from department if not set
        if sample_type == 'Serum':
            dept = safe_str(safe_get(row, tests_headers, ['department']), '').lower()
            if 'urine' in dept:
                sample_type = 'Urine'
            elif 'semen' in dept:
                sample_type = 'Semen'
            elif 'stool' in dept:
                sample_type = 'Stool'
        
        if not test_code or not test_name:
            print(f"  WARNING: Row {row_num}: Missing test_code or test_name, skipping")
            continue
        
        ws_tests.append([
            test_id, test_code, legacy_test_code, test_name,
            category, sample_type, float(price), turnaround_time
        ])
        test_ids.add(test_id)
    
    print(f"  Processed {len(test_ids)} tests")
    
    # ===== 2. PROCESS PARAMETERMASTER -> PARAMETERS SHEET =====
    print("Processing ParameterMaster -> Parameters sheet...")
    param_master_sheet = wb_in['ParameterMaster']
    param_master_headers = get_header_map(param_master_sheet)
    
    ws_params = wb_out.create_sheet("Parameters")
    ws_params.append(['parameter_id', 'parameter_name', 'unit'])
    
    for row_num, row in enumerate(param_master_sheet.iter_rows(min_row=2, values_only=True), 2):
        param_id_raw = safe_get(row, param_master_headers, ['parameter_id', 'param_id', 'id'])
        if not param_id_raw:
            continue
        
        param_id_str = safe_str(param_id_raw)
        # Normalize parameter_id: if numeric, add 'p' prefix
        if param_id_str.isdigit():
            param_id_str = f"p{param_id_str}"
        elif not param_id_str.lower().startswith('p'):
            # Try to extract number and add p prefix
            match = re.search(r'\d+', param_id_str)
            if match:
                param_id_str = f"p{match.group()}"
            else:
                print(f"  WARNING: Row {row_num}: Invalid parameter_id format: {param_id_str}, skipping")
                continue
        
        param_id = param_id_str.lower()
        param_name = safe_str(safe_get(row, param_master_headers, ['parameter_name', 'name']), param_id)
        unit = safe_str(safe_get(row, param_master_headers, ['unit', 'units']))
        
        ws_params.append([param_id, param_name, unit])
        parameter_ids.add(param_id)
    
    print(f"  Processed {len(parameter_ids)} parameters")
    
    # ===== 3. PROCESS PARAMETERS (MAPPING) -> MAPPING SHEET =====
    print("Processing Parameters (mapping) -> Mapping sheet...")
    params_sheet = wb_in['Parameters']
    params_headers = get_header_map(params_sheet)
    
    ws_mapping = wb_out.create_sheet("Mapping")
    ws_mapping.append(['test_id', 'parameter_id', 'display_order', 'reportable'])
    
    for row_num, row in enumerate(params_sheet.iter_rows(min_row=2, values_only=True), 2):
        test_id = safe_int(safe_get(row, params_headers, ['test_id']))
        param_id_raw = safe_get(row, params_headers, ['parameter_id', 'param_id'])
        
        if not test_id or not param_id_raw:
            continue
        
        param_id_str = safe_str(param_id_raw)
        # Normalize parameter_id
        if param_id_str.isdigit():
            param_id_str = f"p{param_id_str}"
        elif not param_id_str.lower().startswith('p'):
            match = re.search(r'\d+', param_id_str)
            if param_id_str and match:
                param_id_str = f"p{match.group()}"
            else:
                continue
        
        param_id = param_id_str.lower()
        display_order = safe_int(safe_get(row, params_headers, ['display_order', 'order']), 0)
        reportable = safe_get(row, params_headers, ['reportable', 'is_active'])
        reportable = True if (reportable is True or str(reportable).lower() in ['true', '1', 'yes', 'y']) else True
        
        # Validate test_id and parameter_id exist
        if test_id not in test_ids:
            print(f"  WARNING: Row {row_num}: test_id {test_id} not found in Tests sheet")
            continue
        if param_id not in parameter_ids:
            print(f"  WARNING: Row {row_num}: parameter_id {param_id} not found in Parameters sheet")
            continue
        
        ws_mapping.append([test_id, param_id, display_order, reportable])
        test_param_mappings.add((test_id, param_id))
    
    print(f"  Processed {len(test_param_mappings)} mappings")
    
    # ===== 4. PROCESS REFERENCERANGES SHEET =====
    print("Processing ReferenceRanges sheet...")
    ws_ranges = wb_out.create_sheet("ReferenceRanges")
    ws_ranges.append([
        'test_id', 'parameter_id', 'gender', 'age_min', 'age_max',
        'reference_min', 'reference_max', 'critical_low', 'critical_high'
    ])
    
    ranges_count = 0
    if 'ReferenceRanges' in wb_in.sheetnames:
        ranges_sheet = wb_in['ReferenceRanges']
        ranges_headers = get_header_map(ranges_sheet)
        
        # Build parameter_name -> parameter_id mapping from Parameters sheet
        param_name_to_id = {}
        for row in params_sheet.iter_rows(min_row=2, values_only=True):
            param_id_raw = safe_get(row, params_headers, ['parameter_id', 'param_id'])
            param_name = safe_str(safe_get(row, params_headers, ['parameter_name', 'name']))
            if param_id_raw and param_name:
                param_id_str = safe_str(param_id_raw)
                if param_id_str.isdigit():
                    param_id_str = f"p{param_id_str}"
                elif not param_id_str.lower().startswith('p'):
                    match = re.search(r'\d+', param_id_str)
                    if match:
                        param_id_str = f"p{match.group()}"
                    else:
                        continue
                param_name_to_id[param_name.lower()] = param_id_str.lower()
        
        for row_num, row in enumerate(ranges_sheet.iter_rows(min_row=2, values_only=True), 2):
            test_id = safe_int(safe_get(row, ranges_headers, ['test_id']))
            param_name = safe_str(safe_get(row, ranges_headers, ['parameter_name', 'name']))
            
            if not test_id:
                continue
            
            # Try to get parameter_id from parameter_name
            param_id = None
            if param_name:
                param_id = param_name_to_id.get(param_name.lower())
            
            # If not found, try direct parameter_id column
            if not param_id:
                param_id_raw = safe_get(row, ranges_headers, ['parameter_id', 'param_id'])
                if param_id_raw:
                    param_id_str = safe_str(param_id_raw)
                    if param_id_str.isdigit():
                        param_id_str = f"p{param_id_str}"
                    elif not param_id_str.lower().startswith('p'):
                        match = re.search(r'\d+', param_id_str)
                        if match:
                            param_id_str = f"p{match.group()}"
                    param_id = param_id_str.lower()
            
            if not param_id:
                continue
            
            gender = safe_str(safe_get(row, ranges_headers, ['gender', 'sex']), 'Both')
            age_min = safe_int(safe_get(row, ranges_headers, ['age_min_years', 'age_min']), 0)
            age_max = safe_int(safe_get(row, ranges_headers, ['age_max_years', 'age_max']), 999)
            ref_min = safe_decimal(safe_get(row, ranges_headers, ['ref_min', 'reference_min']))
            ref_max = safe_decimal(safe_get(row, ranges_headers, ['ref_max', 'reference_max']))
            crit_low = safe_decimal(safe_get(row, ranges_headers, ['critical_low']))
            crit_high = safe_decimal(safe_get(row, ranges_headers, ['critical_high']))
            
            # Validate test_id and parameter_id exist
            if test_id not in test_ids:
                continue
            if param_id not in parameter_ids:
                continue
            
            ws_ranges.append([
                test_id, param_id, gender, age_min, age_max,
                float(ref_min) if ref_min else None,
                float(ref_max) if ref_max else None,
                float(crit_low) if crit_low else None,
                float(crit_high) if crit_high else None
            ])
            ranges_count += 1
    
    print(f"  Processed {ranges_count} reference ranges")
    
    # ===== VALIDATION =====
    print("\n=== Validation Summary ===")
    print(f"Tests: {len(test_ids)}")
    print(f"Parameters: {len(parameter_ids)}")
    print(f"Test-Parameter Mappings: {len(test_param_mappings)}")
    print(f"Reference Ranges: {ranges_count}")
    
    # Check for tests without mappings
    tests_with_mappings = {t_id for t_id, _ in test_param_mappings}
    tests_without_mappings = test_ids - tests_with_mappings
    if tests_without_mappings:
        print(f"\nWARNING: {len(tests_without_mappings)} tests have no parameter mappings")
        print("  These will be handled by catalog_ensure_minimum_parameters command")
    
    # Check for orphaned mappings
    orphaned_mappings = set()
    for test_id, param_id in test_param_mappings:
        if test_id not in test_ids or param_id not in parameter_ids:
            orphaned_mappings.add((test_id, param_id))
    
    if orphaned_mappings:
        print(f"\nERROR: {len(orphaned_mappings)} orphaned mappings found")
        return False
    
    # Save output file
    print(f"\nSaving output to: {output_path}")
    wb_out.save(output_path)
    print("✓ Conversion completed successfully")
    
    return True


def main():
    """Main entry point."""
    if len(sys.argv) < 3:
        print("Usage: convert_excel_to_import_contract.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)
    
    success = convert_excel(input_path, output_path)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
