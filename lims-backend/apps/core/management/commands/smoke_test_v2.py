"""
Run end-to-end API smoke test v2.
"""
import os
import sys
from io import BytesIO
from datetime import datetime

import requests
import openpyxl
from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = "Run end-to-end API smoke test v2"

    def add_arguments(self, parser):
        parser.add_argument("--base-url", default=os.environ.get("BASE_URL", "http://backend:8000"))
        parser.add_argument("--host-header", default=os.environ.get("HOST_HEADER", "localhost"))
        parser.add_argument("--forwarded-proto", default=os.environ.get("FORWARDED_PROTO", "https"))

    def handle(self, *args, **options):
        base_url = options["base_url"].rstrip("/")
        host_header = options["host_header"]
        forwarded_proto = options["forwarded_proto"]
        api_base = f"{base_url}/api/v1"

        username = os.environ.get("ADMIN_USERNAME", "admin")
        password = os.environ.get("ADMIN_PASSWORD", "admin123")

        def fail(step, message):
            self.stdout.write(self.style.ERROR(f"FAIL {step}: {message}"))
            sys.exit(1)

        def ok(step, message):
            self.stdout.write(self.style.SUCCESS(f"PASS {step}: {message}"))

        session = requests.Session()
        session.headers.update({"Host": host_header, "X-Forwarded-Proto": forwarded_proto})

        # Login
        resp = session.post(f"{api_base}/auth/login/", json={"username": username, "password": password})
        if resp.status_code != 200:
            fail("AUTH", f"Login failed ({resp.status_code})")
        data = resp.json()
        token = (data.get("data") or {}).get("access_token") or data.get("access") or data.get("access_token")
        if not token:
            fail("AUTH", "No access token returned")
        ok("AUTH", "Logged in")

        headers = {"Authorization": f"Bearer {token}"}

        # Ensure catalog has a test
        resp = session.get(f"{api_base}/laboratory/tests/", headers=headers)
        if resp.status_code != 200:
            fail("CATALOG-LIST", f"Failed to list tests ({resp.status_code})")
        tests = resp.json().get("results", [])
        if not tests:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            tests_sheet = wb.create_sheet("Tests")
            tests_sheet.append(["test_id", "test_code", "legacy_test_code", "test_name", "category", "sample_type", "sample_volume", "price", "turnaround_time", "loinc_code", "instructions", "is_active"])
            tests_sheet.append([1, "CBC", "", "Complete Blood Count", "Hematology", "Blood", "", 500, 24, "", "", True])
            params_sheet = wb.create_sheet("Parameters")
            params_sheet.append(["parameter_id", "parameter_name", "unit", "data_type", "editor_type", "decimal_places", "allowed_values", "flag_direction", "has_quick_text", "active"])
            params_sheet.append(["p1", "Hemoglobin", "g/dL", "Numeric", "Plain", 2, "", "Both", False, True])
            params_sheet.append(["p2", "WBC", "x10^3/uL", "Numeric", "Plain", 2, "", "Both", False, True])
            mapping_sheet = wb.create_sheet("Mapping")
            mapping_sheet.append(["test_id", "parameter_id", "display_order", "reportable"])
            mapping_sheet.append([1, "p1", 1, True])
            mapping_sheet.append([1, "p2", 2, True])
            ranges_sheet = wb.create_sheet("ReferenceRanges")
            ranges_sheet.append(["test_id", "parameter_id", "gender", "age_min", "age_max", "reference_min", "reference_max", "critical_low", "critical_high", "is_active", "version", "notes"])
            ranges_sheet.append([1, "p1", "Both", 18, 65, 12, 16, 7, 20, True, 1, ""])

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            files = {"file": ("smoke_catalog.xlsx", buffer.getvalue())}
            resp = session.post(
                f"{api_base}/laboratory/import/?strict=true&allow_defaults=true&mode=upsert&dry_run=false",
                headers=headers,
                files=files,
            )
            if resp.status_code not in [200, 201]:
                print(resp.text)
                fail("CATALOG-SEED", f"Failed to import catalog ({resp.status_code})")
            ok("CATALOG-SEED", "Catalog seeded")

            resp = session.get(f"{api_base}/laboratory/tests/", headers=headers)
            tests = resp.json().get("results", [])

        if not tests:
            fail("CATALOG", "No tests available")
        test_id = tests[0].get("test_id") or tests[0].get("id")

        # Create patient
        patient_payload = {
            "first_name": "Smoke",
            "last_name": "Patient",
            "date_of_birth": "1990-01-01",
            "gender": "Male",
            "phone": f"0300{datetime.now().strftime('%H%M%S%f')[:7]}",
        }
        resp = session.post(f"{api_base}/patients/", headers=headers, json=patient_payload)
        if resp.status_code != 201:
            fail("PATIENT", f"Create patient failed ({resp.status_code})")
        patient_id = (resp.json().get("data") or resp.json()).get("id")
        ok("PATIENT", f"Created patient {patient_id}")

        # Create order
        resp = session.post(
            f"{api_base}/orders/orders/",
            headers=headers,
            json={"patient": patient_id, "test_ids": [test_id], "status": "NEW"},
        )
        if resp.status_code != 201:
            fail("ORDER", f"Create order failed ({resp.status_code})")
        order = resp.json()
        order_id = order.get("id")
        items = order.get("items", [])
        if not items:
            fail("ORDER", "No order items created")
        order_item_id = items[0]["id"]
        ok("ORDER", f"Created order {order_id}")

        # Collect sample
        resp = session.get(f"{api_base}/samples/?order_item__order={order_id}", headers=headers)
        if resp.status_code != 200:
            fail("SAMPLES", f"List samples failed ({resp.status_code})")
        samples = resp.json().get("results", [])
        if samples:
            sample_id = samples[0]["id"]
            resp = session.patch(
                f"{api_base}/samples/{sample_id}/",
                headers=headers,
                json={"status": "COLLECTED", "barcode": f"SMOKE-{sample_id}"},
            )
            if resp.status_code != 200:
                fail("SAMPLES", f"Collect sample failed ({resp.status_code})")
        ok("SAMPLES", "Samples collected")

        # Get parameters for test
        resp = session.get(f"{api_base}/laboratory/parameters/?test={test_id}", headers=headers)
        if resp.status_code != 200:
            fail("PARAMETERS", f"List parameters failed ({resp.status_code})")
        params = resp.json().get("results", [])
        if not params:
            fail("PARAMETERS", "No parameters found for test")

        results_payload = {
            "results": [
                {"order_item": order_item_id, "test_parameter": p["id"], "result_value": "1.0"}
                for p in params
            ]
        }
        resp = session.post(f"{api_base}/results/bulk_entry/", headers=headers, json=results_payload)
        if resp.status_code != 201:
            fail("RESULTS", f"Result entry failed ({resp.status_code})")
        ok("RESULTS", "Results entered")

        # Verify results
        resp = session.get(f"{api_base}/results/?order_item={order_item_id}", headers=headers)
        if resp.status_code != 200:
            fail("VERIFY", f"List results failed ({resp.status_code})")
        result_rows = resp.json().get("results", resp.json())
        for row in result_rows:
            verify_resp = session.post(f"{api_base}/results/{row['id']}/verify/", headers=headers)
            if verify_resp.status_code != 200:
                fail("VERIFY", f"Verify failed ({verify_resp.status_code})")
        ok("VERIFY", "Results verified")

        # Update order status to published
        for status in ["COLLECTED", "IN_PROCESS", "VERIFIED", "PUBLISHED"]:
            resp = session.patch(
                f"{api_base}/orders/orders/{order_id}/",
                headers=headers,
                json={"status": status},
            )
            if resp.status_code not in [200, 202]:
                fail("ORDER-STATUS", f"Failed to set {status} ({resp.status_code})")
        ok("ORDER-STATUS", "Order published")

        # Generate report
        resp = session.post(
            f"{api_base}/reports/generate/",
            headers=headers,
            json={"order_id": order_id},
        )
        if resp.status_code not in [200, 201]:
            fail("REPORT", f"Generate report failed ({resp.status_code})")

        resp = session.get(f"{api_base}/orders/orders/{order_id}/report.pdf", headers=headers)
        if resp.status_code != 200:
            fail("REPORT-PDF", f"Download report failed ({resp.status_code})")
        if resp.content[:4] != b"%PDF":
            fail("REPORT-PDF", "Invalid PDF content")
        ok("REPORT-PDF", "Report downloaded")

        # Create payment
        resp = session.post(
            f"{api_base}/payments/",
            headers=headers,
            json={"order": order_id, "amount": order.get("net_amount", "0"), "payment_method": "cash"},
        )
        if resp.status_code != 201:
            fail("PAYMENT", f"Create payment failed ({resp.status_code})")
        payment_id = resp.json().get("id")
        ok("PAYMENT", f"Payment {payment_id}")

        # Receipt PDF
        resp = session.get(f"{api_base}/payments/{payment_id}/receipt/", headers=headers)
        if resp.status_code != 200 or resp.content[:4] != b"%PDF":
            fail("RECEIPT", f"Receipt PDF failed ({resp.status_code})")
        ok("RECEIPT", "Receipt downloaded")

        # Export catalog
        resp = session.get(f"{api_base}/laboratory/export/", headers=headers)
        if resp.status_code != 200:
            fail("CATALOG-EXPORT", f"Export failed ({resp.status_code})")
        export_bytes = resp.content
        ok("CATALOG-EXPORT", "Exported catalog")

        # Import dry run
        files = {"file": ("export.xlsx", export_bytes)}
        resp = session.post(
            f"{api_base}/laboratory/import/?strict=true&allow_defaults=false&mode=upsert&dry_run=true",
            headers=headers,
            files=files,
        )
        if resp.status_code not in [200, 400]:
            fail("CATALOG-IMPORT", f"Dry-run failed ({resp.status_code})")
        summary = resp.json().get("summary", {})
        if summary.get("errors"):
            fail("CATALOG-IMPORT", f"Dry-run errors: {summary.get('errors')[:1]}")
        ok("CATALOG-IMPORT", "Dry-run import ok")

        self.stdout.write(self.style.SUCCESS("Smoke test v2 completed successfully"))
