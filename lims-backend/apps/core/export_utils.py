"""
Utility functions for exporting data to CSV/Excel.
"""

import csv
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def export_to_csv(data, filename="export.csv", headers=None):
    """
    Export data to CSV format.

    Args:
        data: List of dictionaries or list of lists
        filename: Name of the file to download
        headers: Optional list of header names

    Returns:
        HttpResponse: CSV file response
    """
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    if data:
        if isinstance(data[0], dict):
            # Data is list of dictionaries
            if headers is None:
                headers = list(data[0].keys())
            writer.writerow(headers)
            for row in data:
                writer.writerow([row.get(header, "") for header in headers])
        else:
            # Data is list of lists
            if headers:
                writer.writerow(headers)
            for row in data:
                writer.writerow(row)

    return response


def export_to_excel(data, filename="export.xlsx", headers=None, sheet_name="Sheet1"):
    """
    Export data to Excel format.

    Args:
        data: List of dictionaries or list of lists
        filename: Name of the file to download
        headers: Optional list of header names
        sheet_name: Name of the Excel sheet

    Returns:
        HttpResponse: Excel file response
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Style for headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    if data:
        if isinstance(data[0], dict):
            # Data is list of dictionaries
            if headers is None:
                headers = list(data[0].keys())

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = header_alignment

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    # Convert non-serializable types to string
                    if value is not None and not isinstance(
                        value, (str, int, float, bool, type(None))
                    ):
                        try:
                            # Try to convert to string
                            value = str(value)
                        except Exception:
                            # If conversion fails, use empty string
                            value = ""
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Data is list of lists
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                start_row = 2
            else:
                start_row = 1

            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
