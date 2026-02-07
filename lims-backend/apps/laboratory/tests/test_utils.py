"""
Tests for laboratory utility functions.
"""
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.laboratory.models import ReferenceRange, Test, TestCategory, TestParameter
from apps.laboratory.utils import import_tests_from_excel


@pytest.mark.django_db
class TestImportTestsFromExcel:
    """Test import_tests_from_excel utility function."""

    def test_import_tests_sheet(self):
        """Test importing tests from Excel Tests sheet."""
        # Create workbook with Tests sheet
        wb = Workbook()
        ws = wb.create_sheet("Tests")

        # Add header row - updated to match new format
        ws.append(
            [
                "test_id",
                "test_code",
                "legacy_test_code",
                "test_name",
                "category",
                "sample_type",
                "price",
                "turnaround_time",
            ]
        )
        # Add test rows
        ws.append([1, "TEST1", "001", "Test One", "Hematology", "Blood", 100.00, 24])
        ws.append([2, "TEST2", "002", "Test Two", "Chemistry", "Serum", 200.00, 48])

        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Check summary
        assert summary["tests_created"] == 2
        assert Test.objects.filter(test_code="TEST1").exists()
        assert Test.objects.filter(test_code="TEST2").exists()

        # Check categories were created
        assert TestCategory.objects.filter(name="Hematology").exists()
        assert TestCategory.objects.filter(name="Chemistry").exists()

    def test_import_parameters_sheet(self):
        """Test importing parameters from Excel Parameters sheet."""
        from apps.laboratory.models import Parameter

        # Create workbook with Parameters sheet
        wb = Workbook()
        ws = wb.create_sheet("Parameters")

        # Add header row - new format with parameter_id
        ws.append(["parameter_id", "parameter_name", "unit"])
        # Add parameter rows
        ws.append(["p1", "Hemoglobin", "g/dL"])
        ws.append(["p2", "WBC", "10^3/uL"])

        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Check summary
        assert summary["parameters_created"] == 2
        assert Parameter.objects.filter(parameter_id="p1").exists()
        assert Parameter.objects.filter(parameter_id="p2").exists()

    def test_import_reference_ranges_sheet(self):
        """Test importing reference ranges from Excel ReferenceRanges sheet."""
        from apps.laboratory.models import Parameter

        # Create test and parameter first
        category = TestCategory.objects.create(name="Hematology")
        test = Test.objects.create(
            test_id=1,
            category=category,
            test_code="TEST1",
            test_name="Test One",
            sample_type="Blood",
            price=Decimal("100.00"),
            turnaround_time=24,
        )
        parameter = Parameter.objects.create(
            parameter_id="p1",
            parameter_name="Hemoglobin",
            unit="g/dL",
        )
        test_param = TestParameter.objects.create(
            test=test,
            parameter=parameter,
        )

        # Create workbook with ReferenceRanges sheet
        wb = Workbook()
        ws = wb.create_sheet("ReferenceRanges")

        # Add header row - new format with test_id and parameter_id
        ws.append(
            [
                "test_id",
                "parameter_id",
                "gender",
                "age_min",
                "age_max",
                "reference_min",
                "reference_max",
                "critical_low",
                "critical_high",
            ]
        )
        # Add range row
        ws.append([1, "p1", "Male", 18, 65, 10.0, 20.0, 5.0, 25.0])

        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Check summary
        assert summary["ranges_created"] == 1
        assert ReferenceRange.objects.filter(
            parameter=test_param,
            gender="Male",
            age_min=18,
            age_max=65,
        ).exists()

    def test_import_all_sheets_together(self):
        """Test importing all sheets together."""
        from apps.laboratory.models import Parameter

        wb = Workbook()

        # Parameters sheet (must come first to define parameters)
        ws_params = wb.create_sheet("Parameters")
        ws_params.append(["parameter_id", "parameter_name", "unit"])
        ws_params.append(["p1", "Hemoglobin", "g/dL"])

        # Tests sheet
        ws_tests = wb.create_sheet("Tests")
        ws_tests.append(
            [
                "test_id",
                "test_code",
                "legacy_test_code",
                "test_name",
                "category",
                "sample_type",
                "price",
                "turnaround_time",
            ]
        )
        ws_tests.append(
            [1, "TEST1", "001", "Test One", "Hematology", "Blood", 100.00, 24]
        )

        # Mapping sheet (links tests to parameters)
        ws_mapping = wb.create_sheet("Mapping")
        ws_mapping.append(["test_id", "parameter_id", "display_order", "reportable"])
        ws_mapping.append([1, "p1", 1, True])

        # ReferenceRanges sheet
        ws_ranges = wb.create_sheet("ReferenceRanges")
        ws_ranges.append(
            [
                "test_id",
                "parameter_id",
                "gender",
                "age_min",
                "age_max",
                "reference_min",
                "reference_max",
                "critical_low",
                "critical_high",
            ]
        )
        ws_ranges.append([1, "p1", "Male", 18, 65, 10.0, 20.0, 5.0, 25.0])

        # Save to BytesIO
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Check all data imported
        assert summary["tests_created"] == 1
        assert summary["parameters_created"] == 1
        assert summary["mappings_created"] == 1
        assert summary["ranges_created"] == 1

        test = Test.objects.get(test_code="TEST1")
        parameter = Parameter.objects.get(parameter_id="p1")
        test_param = TestParameter.objects.get(test=test, parameter=parameter)
        assert ReferenceRange.objects.filter(parameter=test_param).exists()

    def test_import_updates_existing_test(self):
        """Test that import updates existing test instead of creating duplicate."""
        category = TestCategory.objects.create(name="Hematology")
        test = Test.objects.create(
            test_id=1,
            category=category,
            test_code="TEST1",
            test_name="Old Name",
            sample_type="Blood",
            price=Decimal("50.00"),
            turnaround_time=12,
        )

        # Create workbook with updated test data
        wb = Workbook()
        ws = wb.create_sheet("Tests")
        ws.append(
            [
                "test_id",
                "test_code",
                "legacy_test_code",
                "test_name",
                "category",
                "sample_type",
                "price",
                "turnaround_time",
            ]
        )
        ws.append([1, "TEST1", "001", "New Name", "Hematology", "Serum", 100.00, 24])

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Check test was updated, not created
        assert summary["tests_updated"] == 1
        assert summary["tests_created"] == 0

        test.refresh_from_db()
        assert test.test_name == "New Name"
        assert test.sample_type == "Serum"

    def test_import_skips_empty_rows(self):
        """Test that import skips empty rows."""
        wb = Workbook()
        ws = wb.create_sheet("Tests")
        ws.append(
            [
                "test_id",
                "test_code",
                "legacy_test_code",
                "test_name",
                "category",
                "sample_type",
                "price",
                "turnaround_time",
            ]
        )
        ws.append([1, "TEST1", "001", "Test One", "Hematology", "Blood", 100.00, 24])
        ws.append([None, None, None, None, None, None, None, None])  # Empty row
        ws.append([2, "TEST2", "002", "Test Two", "Chemistry", "Serum", 200.00, 48])

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import
        summary = import_tests_from_excel(file_obj)

        # Should only create 2 tests, skipping empty row
        assert summary["tests_created"] == 2

    def test_import_handles_missing_test_for_mapping(self):
        """Test that import handles missing test gracefully for mapping."""
        from apps.laboratory.models import Parameter

        # Create a parameter
        Parameter.objects.create(parameter_id="p1", parameter_name="Hemoglobin")

        wb = Workbook()
        ws = wb.create_sheet("Mapping")
        ws.append(["test_id", "parameter_id", "display_order", "reportable"])
        ws.append([999, "p1", 1, True])  # test_id 999 doesn't exist

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import should not raise error but should log error
        summary = import_tests_from_excel(file_obj)

        # No mappings should be created
        assert summary["mappings_created"] == 0
        assert len(summary["errors"]) > 0

    def test_import_handles_missing_parameter_for_range(self):
        """Test that import handles missing parameter gracefully for ranges."""
        category = TestCategory.objects.create(name="Hematology")
        test = Test.objects.create(
            test_id=1,
            category=category,
            test_code="TEST1",
            test_name="Test One",
            sample_type="Blood",
            price=Decimal("100.00"),
            turnaround_time=24,
        )

        wb = Workbook()
        ws = wb.create_sheet("ReferenceRanges")
        ws.append(
            [
                "test_id",
                "parameter_id",
                "gender",
                "age_min",
                "age_max",
                "reference_min",
                "reference_max",
                "critical_low",
                "critical_high",
            ]
        )
        ws.append(
            [1, "p999", "Male", 18, 65, 10.0, 20.0, 5.0, 25.0]
        )  # p999 doesn't exist

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        # Import should not raise error but should log error
        summary = import_tests_from_excel(file_obj)

        # No ranges should be created
        assert summary["ranges_created"] == 0
        assert len(summary["errors"]) > 0
