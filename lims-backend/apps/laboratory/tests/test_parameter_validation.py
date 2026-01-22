"""
Tests for parameter_id validation.
"""
import pytest
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from django.core.exceptions import ValidationError
from apps.laboratory.models import (
    Parameter, 
    Test, 
    TestCategory, 
    TestParameter,
    validate_parameter_id
)
from apps.laboratory.utils import import_tests_from_excel


@pytest.mark.django_db
class TestParameterIdValidation:
    """Test parameter_id format validation."""
    
    def test_valid_parameter_id_formats(self):
        """Test that valid parameter_id formats are accepted."""
        valid_ids = ["p1", "p2", "p10", "p53", "p100", "p999"]
        
        for param_id in valid_ids:
            param = Parameter(
                parameter_id=param_id,
                parameter_name="Test Parameter"
            )
            param.save()
            
            # Verify it was saved with lowercase
            assert Parameter.objects.filter(parameter_id=param_id.lower()).exists()
    
    def test_uppercase_parameter_id_normalized(self):
        """Test that uppercase parameter_ids are normalized to lowercase."""
        param = Parameter(
            parameter_id="P1",
            parameter_name="Test Parameter"
        )
        param.save()
        
        # Should be stored as lowercase
        assert param.parameter_id == "p1"
        assert Parameter.objects.filter(parameter_id="p1").exists()
    
    def test_invalid_parameter_id_formats(self):
        """Test that invalid parameter_id formats are rejected."""
        invalid_ids = [
            "P1x",      # Has letters after number
            "1",        # Missing 'p' prefix
            "param1",   # Wrong prefix
            "",         # Empty
            "p",        # Missing number
            "pp1",      # Double prefix
            "p1a",      # Letter after number
        ]
        
        for param_id in invalid_ids:
            with pytest.raises(ValidationError):
                param = Parameter(
                    parameter_id=param_id,
                    parameter_name="Test Parameter"
                )
                param.save()
    
    def test_validate_parameter_id_function(self):
        """Test the validate_parameter_id function directly."""
        # Valid cases
        assert validate_parameter_id("p1") == "p1"
        assert validate_parameter_id("P1") == "p1"
        assert validate_parameter_id("p100") == "p100"
        assert validate_parameter_id(" p5 ") == "p5"  # Strips whitespace
        
        # Invalid cases
        with pytest.raises(ValidationError):
            validate_parameter_id("")
        
        with pytest.raises(ValidationError):
            validate_parameter_id("param1")
        
        with pytest.raises(ValidationError):
            validate_parameter_id("p1x")
    
    def test_parameter_id_uniqueness(self):
        """Test that parameter_id must be unique."""
        Parameter.objects.create(
            parameter_id="p1",
            parameter_name="First Parameter"
        )
        
        # Trying to create another with same ID should fail
        with pytest.raises(Exception):  # IntegrityError
            Parameter.objects.create(
                parameter_id="p1",
                parameter_name="Second Parameter"
            )
    
    def test_parameter_id_case_insensitive_uniqueness(self):
        """Test that parameter_id uniqueness is case-insensitive."""
        Parameter.objects.create(
            parameter_id="p1",
            parameter_name="First Parameter"
        )
        
        # Trying to create with uppercase should fail (normalized to same ID)
        with pytest.raises(Exception):  # IntegrityError
            Parameter.objects.create(
                parameter_id="P1",
                parameter_name="Second Parameter"
            )


@pytest.mark.django_db
class TestExcelImportParameterValidation:
    """Test parameter_id validation during Excel import."""
    
    def test_import_valid_parameter_ids(self):
        """Test importing parameters with valid parameter_ids."""
        wb = Workbook()
        ws = wb.create_sheet("Parameters")
        
        ws.append(["parameter_id", "parameter_name", "unit"])
        ws.append(["p1", "Hemoglobin", "g/dL"])
        ws.append(["p2", "WBC", "10^3/uL"])
        ws.append(["p53", "Glucose", "mg/dL"])
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj)
        
        assert summary["parameters_created"] == 3
        assert len(summary["errors"]) == 0
        assert Parameter.objects.filter(parameter_id="p1").exists()
        assert Parameter.objects.filter(parameter_id="p2").exists()
        assert Parameter.objects.filter(parameter_id="p53").exists()
    
    def test_import_invalid_parameter_id_format(self):
        """Test that import rejects invalid parameter_id formats."""
        wb = Workbook()
        ws = wb.create_sheet("Parameters")
        
        ws.append(["parameter_id", "parameter_name", "unit"])
        ws.append(["param1", "Invalid Param", "g/dL"])  # Invalid format
        ws.append(["p2", "Valid Param", "mg/dL"])       # Valid
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj)
        
        # Should create only the valid one
        assert summary["parameters_created"] == 1
        assert len(summary["errors"]) == 1
        assert Parameter.objects.filter(parameter_id="p2").exists()
        assert not Parameter.objects.filter(parameter_id="param1").exists()
        
        # Check error message
        error = summary["errors"][0]
        assert error["sheet"] == "Parameters"
        assert error["column"] == "parameter_id"
        assert "format" in error["message"].lower()
    
    def test_import_duplicate_parameter_ids(self):
        """Test that import detects duplicate parameter_ids."""
        wb = Workbook()
        ws = wb.create_sheet("Parameters")
        
        ws.append(["parameter_id", "parameter_name", "unit"])
        ws.append(["p1", "First", "g/dL"])
        ws.append(["p1", "Duplicate", "mg/dL"])  # Duplicate
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj)
        
        # Should create only one
        assert summary["parameters_created"] == 1
        assert len(summary["errors"]) == 1
        
        # Check error message
        error = summary["errors"][0]
        assert error["sheet"] == "Parameters"
        assert "duplicate" in error["message"].lower()
    
    def test_import_mapping_with_missing_parameter_id(self):
        """Test that import fails when mapping references non-existent parameter."""
        wb = Workbook()
        
        # Create test
        category = TestCategory.objects.create(name="Hematology")
        test = Test.objects.create(
            test_id=1,
            test_code="TEST1",
            test_name="Test One",
            category=category,
            sample_type="Blood",
            price=Decimal("100.00"),
            turnaround_time=24,
        )
        
        # Create mapping that references non-existent parameter
        ws_mapping = wb.create_sheet("Mapping")
        ws_mapping.append(["test_id", "parameter_id", "display_order", "reportable"])
        ws_mapping.append([1, "p999", 1, True])  # p999 doesn't exist
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj)
        
        # Should have error about missing parameter
        assert len(summary["errors"]) > 0
        error = summary["errors"][0]
        assert error["sheet"] == "Mapping"
        assert "p999" in error["message"]
    
    def test_import_mapping_with_invalid_parameter_id_format(self):
        """Test that mapping sheet validates parameter_id format."""
        wb = Workbook()
        
        # Create test
        category = TestCategory.objects.create(name="Hematology")
        test = Test.objects.create(
            test_id=1,
            test_code="TEST1",
            test_name="Test One",
            category=category,
            sample_type="Blood",
            price=Decimal("100.00"),
            turnaround_time=24,
        )
        
        # Create mapping with invalid parameter_id format
        ws_mapping = wb.create_sheet("Mapping")
        ws_mapping.append(["test_id", "parameter_id", "display_order", "reportable"])
        ws_mapping.append([1, "param1", 1, True])  # Invalid format
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj)
        
        # Should have error about invalid format
        assert len(summary["errors"]) > 0
        error = summary["errors"][0]
        assert error["sheet"] == "Mapping"
        assert "format" in error["message"].lower()


@pytest.mark.django_db
class TestDryRunImport:
    """Test dry-run functionality."""
    
    def test_dry_run_validates_without_writing(self):
        """Test that dry-run validates but doesn't write to database."""
        initial_count = Parameter.objects.count()
        
        wb = Workbook()
        ws = wb.create_sheet("Parameters")
        
        ws.append(["parameter_id", "parameter_name", "unit"])
        ws.append(["p1", "Hemoglobin", "g/dL"])
        ws.append(["p2", "WBC", "10^3/uL"])
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj, dry_run=True)
        
        # Should indicate what would be created
        assert summary["parameters_created"] == 2
        assert summary["dry_run"] is True
        assert summary["status"] == "PASS"
        
        # But database should be unchanged
        assert Parameter.objects.count() == initial_count
    
    def test_dry_run_detects_errors(self):
        """Test that dry-run detects validation errors."""
        wb = Workbook()
        ws = wb.create_sheet("Parameters")
        
        ws.append(["parameter_id", "parameter_name", "unit"])
        ws.append(["invalid", "Bad Param", "g/dL"])  # Invalid format
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        summary = import_tests_from_excel(file_obj, dry_run=True)
        
        # Should detect error
        assert summary["dry_run"] is True
        assert summary["status"] == "FAIL"
        assert len(summary["errors"]) > 0
        assert not summary["validation_passed"]
    
    def test_dry_run_full_workflow(self):
        """Test dry-run with complete workflow: Tests, Parameters, Mapping."""
        wb = Workbook()
        
        # Parameters sheet
        ws_params = wb.create_sheet("Parameters")
        ws_params.append(["parameter_id", "parameter_name", "unit"])
        ws_params.append(["p1", "Hemoglobin", "g/dL"])
        
        # Tests sheet
        ws_tests = wb.create_sheet("Tests")
        ws_tests.append(["test_id", "test_code", "legacy_test_code", "test_name", 
                        "category", "sample_type", "price", "turnaround_time"])
        ws_tests.append([1, "HB", "001", "Hemoglobin Test", "Hematology", 
                        "Blood", 50.00, 24])
        
        # Mapping sheet
        ws_mapping = wb.create_sheet("Mapping")
        ws_mapping.append(["test_id", "parameter_id", "display_order", "reportable"])
        ws_mapping.append([1, "p1", 1, True])
        
        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)
        
        initial_param_count = Parameter.objects.count()
        initial_test_count = Test.objects.count()
        initial_mapping_count = TestParameter.objects.count()
        
        summary = import_tests_from_excel(file_obj, dry_run=True)
        
        # Should pass validation
        assert summary["status"] == "PASS"
        assert summary["validation_passed"]
        assert len(summary["errors"]) == 0
        
        # But database should be unchanged
        assert Parameter.objects.count() == initial_param_count
        assert Test.objects.count() == initial_test_count
        assert TestParameter.objects.count() == initial_mapping_count
