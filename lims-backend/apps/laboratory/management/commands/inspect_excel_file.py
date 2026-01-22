from django.core.management.base import BaseCommand
import openpyxl
import os

class Command(BaseCommand):
    help = 'Inspect Excel file headers'

    def handle(self, *args, **options):
        file_path = "LIMS_TestCatalog_MVP_FINAL (1).xlsx"
        if not os.path.exists(file_path): # Check if file exists in root or where we run it
             # Try one level up if not found, since we are in lims-backend usually
             if os.path.exists(f"../{file_path}"):
                 file_path = f"../{file_path}"
             else:
                 self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
                 return

        try:
            wb = openpyxl.load_workbook(file_path)
            self.stdout.write(f"File: {file_path}")
            self.stdout.write(f"Sheet names: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = [str(cell.value).strip() if cell.value else None for cell in sheet[1]]
                self.stdout.write(f"\n--- Sheet: {sheet_name} ---")
                self.stdout.write(f"Columns: {headers}")
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error reading excel: {e}"))
