"""
Catalog import/export utilities with strict validation and deterministic output.
"""
from __future__ import annotations

from copy import deepcopy
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from django.db import transaction

from .models import (
    Parameter,
    ReferenceRange,
    Test,
    TestCategory,
    TestPanel,
    TestParameter,
    validate_parameter_id,
)

SHEET_ORDER = [
    "Tests",
    "Parameters",
    "Mapping",
    "Panels",
    "PanelTests",
    "ReferenceRanges",
]

# Column aliases - map common variations to expected column names
# Format: {normalized_alias: normalized_expected_name}
COLUMN_ALIASES = {
    # Tests sheet aliases
    "tat_hours": "turnaround_time",
    "tat": "turnaround_time",
    "turnaround_time_hours": "turnaround_time",
    "sample_volume_ml": "sample_volume",
    "volume_ml": "sample_volume",
    "test_category": "category",
    "name": "test_name",
    # Parameters sheet aliases
    "field_type": "data_type",
    "type": "data_type",
    "param_id": "parameter_id",
    "param_name": "parameter_name",
    "options": "allowed_values",
    # ReferenceRanges sheet aliases
    "age_min_years": "age_min",
    "age_max_years": "age_max",
    "min_age": "age_min",
    "max_age": "age_max",
    "ref_min": "reference_min",
    "ref_max": "reference_max",
    "min_value": "reference_min",
    "max_value": "reference_max",
    "low": "reference_min",
    "high": "reference_max",
    "crit_low": "critical_low",
    "crit_high": "critical_high",
    # Panels sheet aliases
    "panel_id": "panel_code",
    "name": "panel_name",
    "panel_category": "category",
}

# Values that should be treated as None/null
NULL_VALUES = frozenset(
    [
        "na",
        "n/a",
        "null",
        "none",
        "-",
        "--",
        ".",
        "",
        "#n/a",
        "#null",
        "#na",
        "nil",
        "undefined",
    ]
)

CATALOG_COLUMNS = {
    "Tests": [
        "test_id",
        "test_code",
        "legacy_test_code",
        "test_name",
        "category",
        "sample_type",
        "sample_volume",
        "price",
        "turnaround_time",
        "loinc_code",
        "instructions",
        "is_active",
    ],
    "Parameters": [
        "parameter_id",
        "parameter_name",
        "unit",
        "data_type",
        "editor_type",
        "decimal_places",
        "allowed_values",
        "flag_direction",
        "has_quick_text",
        "active",
    ],
    "Mapping": [
        "test_id",
        "parameter_id",
        "display_order",
        "reportable",
    ],
    "Panels": [
        "panel_code",
        "panel_name",
        "category",
        "sample_type",
        "sample_volume",
        "price",
        "turnaround_time",
        "description",
        "is_active",
    ],
    "PanelTests": [
        "panel_code",
        "test_id",
    ],
    "ReferenceRanges": [
        "test_id",
        "parameter_id",
        "gender",
        "age_min",
        "age_max",
        "reference_min",
        "reference_max",
        "critical_low",
        "critical_high",
        "is_active",
        "version",
        "notes",
    ],
}

REQUIRED_FIELDS = {
    "Tests": ["test_id", "test_code", "test_name", "category"],
    "Parameters": ["parameter_id", "parameter_name"],
    "Mapping": ["test_id", "parameter_id"],
    "Panels": ["panel_code", "panel_name", "category"],
    "PanelTests": ["panel_code", "test_id"],
    "ReferenceRanges": ["test_id", "parameter_id"],
}

DEFAULTS = {
    "Tests": {
        "sample_type": "Serum",
        "price": Decimal("0"),
        "turnaround_time": 24,
        "is_active": True,
    },
    "Parameters": {
        "data_type": "Numeric",
        "editor_type": "Plain",
        "decimal_places": 2,
        "flag_direction": "Both",
        "has_quick_text": False,
        "active": True,
    },
    "Mapping": {
        "display_order": 0,
        "reportable": True,
    },
    "Panels": {
        "sample_type": "Serum",
        "price": Decimal("0"),
        "turnaround_time": 24,
        "is_active": True,
    },
    "ReferenceRanges": {
        "gender": "Both",
        "is_active": True,
        "version": 1,
    },
}


def normalize_header(value: Any) -> str:
    """Normalize a header value to lowercase with underscores."""
    return (
        str(value).strip().lower().replace(" ", "_").replace("(", "").replace(")", "")
    )


def apply_column_aliases(headers: Dict[str, int]) -> Dict[str, int]:
    """Apply column aliases to headers, allowing common variations."""
    result = dict(headers)
    for alias, canonical in COLUMN_ALIASES.items():
        if alias in headers and canonical not in headers:
            result[canonical] = headers[alias]
    return result


def get_headers(sheet, apply_aliases: bool = True) -> Dict[str, int]:
    """
    Extract headers from the first row of a sheet.

    Args:
        sheet: The openpyxl worksheet
        apply_aliases: If True, apply column name aliases for compatibility

    Returns:
        Dictionary mapping normalized header names to column indices
    """
    headers: Dict[str, int] = {}
    if sheet and sheet.max_row >= 1:
        for idx, cell in enumerate(sheet[1]):
            if cell.value is None:
                continue
            headers[normalize_header(cell.value)] = idx

    if apply_aliases:
        headers = apply_column_aliases(headers)

    return headers


def is_null_value(value: Any) -> bool:
    """Check if a value should be treated as null/None."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in NULL_VALUES
    return False


def safe_get(row: Tuple[Any, ...], headers: Dict[str, int], key: str) -> Any:
    """Get a value from a row, returning None if not found or if it's a null-like value."""
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    if is_null_value(value):
        return None
    return value


def to_decimal(value: Any) -> Optional[Decimal]:
    """Convert a value to Decimal, treating null-like values as None."""
    if is_null_value(value):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def to_int(value: Any) -> Optional[int]:
    """Convert a value to int, treating null-like values as None."""
    if is_null_value(value):
        return None
    try:
        # Handle floats by converting to int (e.g., 24.0 -> 24)
        val_str = str(value).strip()
        if "." in val_str:
            return int(float(val_str))
        return int(val_str)
    except (TypeError, ValueError):
        return None


def to_bool(value: Any, default: Optional[bool] = None) -> Optional[bool]:
    """Convert a value to bool, treating null-like values as the default."""
    if is_null_value(value):
        return default
    if isinstance(value, bool):
        return value
    value_str = str(value).strip().lower()
    if value_str in ["true", "1", "yes", "y"]:
        return True
    if value_str in ["false", "0", "no", "n"]:
        return False
    return default


def deep_merge(
    base: Dict[str, Any], override: Optional[Dict[str, Any]]
) -> Dict[str, Any]:
    merged = deepcopy(base)
    if not override:
        return merged
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = deep_merge(merged[key], value)
        else:
            merged[key] = value
    return merged


def _add_issue(
    collector: List[Dict[str, Any]], sheet: str, row: int, field: str, message: str
):
    collector.append({"sheet": sheet, "row": row, "field": field, "message": message})


def _record_diff(
    diff: List[Dict[str, Any]],
    sheet: str,
    key: str,
    action: str,
    changes: Dict[str, Any],
):
    diff.append(
        {
            "sheet": sheet,
            "key": key,
            "action": action,
            "changes": changes,
        }
    )


def _normalize_param_id(value: Any) -> Optional[str]:
    if is_null_value(value):
        return None
    p_id_str = str(value).strip()
    if p_id_str.isdigit():
        p_id_str = f"p{p_id_str}"
    return validate_parameter_id(p_id_str)


def _serialize_for_json(obj: Any) -> Any:
    """Recursively convert Decimal and other non-JSON types to JSON-safe values."""
    if isinstance(obj, Decimal):
        return str(obj)
    if isinstance(obj, dict):
        return {k: _serialize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [_serialize_for_json(v) for v in obj]
    return obj


def _validate_sheet_headers(
    sheet_name: str,
    headers: Dict[str, int],
    required_fields: List[str],
    warnings: List[Dict[str, Any]],
) -> List[str]:
    """
    Validate that required headers are present and return list of missing ones.
    Also adds warnings for unrecognized headers.
    """
    missing = []
    for field in required_fields:
        if field not in headers:
            missing.append(field)

    # Check for unknown headers (after alias application)
    expected_columns = set(CATALOG_COLUMNS.get(sheet_name, []))
    found_headers = set(headers.keys())
    unknown = found_headers - expected_columns

    if unknown:
        warnings.append(
            {
                "sheet": sheet_name,
                "row": 1,
                "field": "headers",
                "message": f"Unrecognized columns (ignored): {', '.join(sorted(unknown))}",
            }
        )

    return missing


def import_catalog_from_excel(
    file,
    *,
    strict: bool = True,
    allow_defaults: bool = False,
    mode: str = "upsert",
    dry_run: bool = True,
) -> Dict[str, Any]:
    if mode != "upsert":
        raise ValueError("Only mode='upsert' is supported")

    workbook = openpyxl.load_workbook(file)
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    diff: List[Dict[str, Any]] = []

    summary = {
        "tests": {"created": 0, "updated": 0, "unchanged": 0},
        "parameters": {"created": 0, "updated": 0, "unchanged": 0},
        "mappings": {"created": 0, "updated": 0, "unchanged": 0},
        "panels": {"created": 0, "updated": 0, "unchanged": 0},
        "panel_tests": {"created": 0, "updated": 0, "unchanged": 0},
        "reference_ranges": {"created": 0, "updated": 0, "unchanged": 0},
    }

    existing_tests = {
        t.test_id: t for t in Test.objects.select_related("category").all()
    }
    existing_params = {p.parameter_id: p for p in Parameter.objects.all()}
    existing_panels = {
        p.panel_code: p for p in TestPanel.objects.select_related("category").all()
    }
    existing_mappings = {
        (tp.test_id, tp.parameter_id): tp
        for tp in TestParameter.objects.select_related("test", "parameter").all()
    }
    existing_ranges = {
        (
            rr.parameter.test_id,
            rr.parameter.parameter_id,
            rr.gender,
            rr.age_min,
            rr.age_max,
            rr.version,
        ): rr
        for rr in ReferenceRange.objects.select_related("parameter").all()
    }

    created_tests: Dict[int, Test] = {}
    created_params: Dict[str, Parameter] = {}
    created_panels: Dict[str, TestPanel] = {}
    created_mappings: Dict[Tuple[int, str], TestParameter] = {}

    def require_value(sheet, row_num, field, value):
        if value is None or str(value).strip() == "":
            if strict:
                _add_issue(errors, sheet, row_num, field, "Missing required value")
            else:
                _add_issue(warnings, sheet, row_num, field, "Missing required value")
            return False
        return True

    def apply_default(sheet, row_num, field, value, default):
        if value is None or str(value).strip() == "":
            if allow_defaults:
                _add_issue(warnings, sheet, row_num, field, f"Defaulted to {default!r}")
                return default
            if strict:
                _add_issue(
                    errors,
                    sheet,
                    row_num,
                    field,
                    "Missing required value (defaults disabled)",
                )
            else:
                _add_issue(
                    warnings,
                    sheet,
                    row_num,
                    field,
                    "Missing required value (defaults disabled)",
                )
            return None
        return value

    def compare_fields(
        existing_obj, incoming: Dict[str, Any], fields: Iterable[str], provided_fields: set = None
    ) -> Dict[str, Any]:
        """
        Compare fields between existing object and incoming data.
        Only compares fields that were explicitly provided (not None/null).
        
        Args:
            existing_obj: The existing database object
            incoming: Dictionary of incoming values
            fields: List of field names to compare
            provided_fields: Set of field names that were explicitly provided (not None/null)
        """
        if provided_fields is None:
            provided_fields = set(fields)
        
        changes: Dict[str, Any] = {}
        for field in fields:
            # Only compare fields that were explicitly provided
            if field not in provided_fields:
                continue
                
            current = getattr(existing_obj, field)
            incoming_val = incoming.get(field)

            current_for_diff = str(current) if isinstance(current, Decimal) else current
            incoming_for_diff = (
                str(incoming_val) if isinstance(incoming_val, Decimal) else incoming_val
            )

            if str(current) != str(incoming_val):
                changes[field] = {"from": current_for_diff, "to": incoming_for_diff}
        return changes

    transaction_context = transaction.atomic() if not dry_run else DummyContext()

    with transaction_context:
        # Tests
        if "Tests" in workbook.sheetnames:
            sheet = workbook["Tests"]
            headers = get_headers(sheet)
            seen_ids = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                test_id = to_int(safe_get(row, headers, "test_id"))
                if test_id is None:
                    continue
                if test_id in seen_ids:
                    _add_issue(
                        errors,
                        "Tests",
                        row_num,
                        "test_id",
                        f"Duplicate test_id {test_id}",
                    )
                    continue
                seen_ids.add(test_id)

                test_code = safe_get(row, headers, "test_code")
                test_name = safe_get(row, headers, "test_name")
                category_name = safe_get(row, headers, "category")

                if not require_value("Tests", row_num, "test_code", test_code):
                    continue
                if not require_value("Tests", row_num, "test_name", test_name):
                    continue
                if not require_value("Tests", row_num, "category", category_name):
                    continue

                # Track which fields were explicitly provided (not None/null)
                provided_fields = {"test_code", "test_name", "category_name"}
                
                # For optional fields, check if they were provided
                raw_sample_type = safe_get(row, headers, "sample_type")
                raw_price = safe_get(row, headers, "price")
                raw_tat = safe_get(row, headers, "turnaround_time")
                raw_is_active = safe_get(row, headers, "is_active")
                raw_legacy_code = safe_get(row, headers, "legacy_test_code")
                raw_sample_volume = safe_get(row, headers, "sample_volume")
                raw_loinc_code = safe_get(row, headers, "loinc_code")
                raw_instructions = safe_get(row, headers, "instructions")

                # Apply defaults only if allow_defaults is True, otherwise use None
                if raw_sample_type is not None:
                    sample_type = str(raw_sample_type).strip()
                    provided_fields.add("sample_type")
                elif allow_defaults:
                    sample_type = DEFAULTS["Tests"]["sample_type"]
                    _add_issue(warnings, "Tests", row_num, "sample_type", f"Defaulted to {sample_type!r}")
                else:
                    sample_type = None

                if raw_price is not None:
                    price_val = to_decimal(raw_price)
                    if price_val is not None:
                        provided_fields.add("price")
                elif allow_defaults:
                    price_val = DEFAULTS["Tests"]["price"]
                    _add_issue(warnings, "Tests", row_num, "price", f"Defaulted to {price_val!r}")
                else:
                    price_val = None

                if raw_tat is not None:
                    tat_val = to_int(raw_tat)
                    if tat_val is not None:
                        provided_fields.add("turnaround_time")
                elif allow_defaults:
                    tat_val = DEFAULTS["Tests"]["turnaround_time"]
                    _add_issue(warnings, "Tests", row_num, "turnaround_time", f"Defaulted to {tat_val!r}")
                else:
                    tat_val = None

                # For required fields with defaults, ensure we have values
                if sample_type is None or price_val is None or tat_val is None:
                    if strict:
                        if sample_type is None:
                            _add_issue(errors, "Tests", row_num, "sample_type", "Missing required value")
                        if price_val is None:
                            _add_issue(errors, "Tests", row_num, "price", "Missing required value")
                        if tat_val is None:
                            _add_issue(errors, "Tests", row_num, "turnaround_time", "Missing required value")
                    continue

                is_active_val = None
                if raw_is_active is not None:
                    is_active_val = to_bool(raw_is_active, DEFAULTS["Tests"]["is_active"])
                    provided_fields.add("is_active")
                elif allow_defaults:
                    is_active_val = DEFAULTS["Tests"]["is_active"]
                    _add_issue(warnings, "Tests", row_num, "is_active", f"Defaulted to {is_active_val!r}")
                else:
                    is_active_val = DEFAULTS["Tests"]["is_active"]  # Use default for new records

                incoming = {
                    "test_code": str(test_code).strip(),
                    "legacy_test_code": (
                        str(raw_legacy_code).strip() or None
                    )
                    if raw_legacy_code is not None
                    else None,
                    "test_name": str(test_name).strip(),
                    "category_name": str(category_name).strip(),
                    "sample_type": str(sample_type).strip(),
                    "sample_volume": (
                        str(raw_sample_volume).strip() or None
                    )
                    if raw_sample_volume is not None
                    else None,
                    "price": price_val,
                    "turnaround_time": int(tat_val),
                    "loinc_code": (
                        str(raw_loinc_code).strip() or None
                    )
                    if raw_loinc_code is not None
                    else None,
                    "instructions": (
                        str(raw_instructions).strip() or None
                    )
                    if raw_instructions is not None
                    else None,
                    "is_active": is_active_val if is_active_val is not None else True,
                }

                # Track optional fields that were provided
                if raw_legacy_code is not None:
                    provided_fields.add("legacy_test_code")
                if raw_sample_volume is not None:
                    provided_fields.add("sample_volume")
                if raw_loinc_code is not None:
                    provided_fields.add("loinc_code")
                if raw_instructions is not None:
                    provided_fields.add("instructions")

                existing = existing_tests.get(test_id)
                if existing:
                    # Only compare and update fields that were explicitly provided
                    fields_to_compare = [
                        "test_code",
                        "legacy_test_code",
                        "test_name",
                        "sample_type",
                        "sample_volume",
                        "price",
                        "turnaround_time",
                        "loinc_code",
                        "instructions",
                        "is_active",
                    ]
                    changes = compare_fields(
                        existing,
                        incoming,
                        fields_to_compare,
                        provided_fields=provided_fields,
                    )
                    if "category_name" in provided_fields and existing.category.name != incoming["category_name"]:
                        changes["category"] = {
                            "from": existing.category.name
                            if existing.category
                            else None,
                            "to": incoming["category_name"],
                        }

                    if changes:
                        summary["tests"]["updated"] += 1
                        _record_diff(diff, "Tests", str(test_id), "update", changes)
                        if not dry_run:
                            if "category_name" in provided_fields:
                                category, _ = TestCategory.objects.get_or_create(
                                    name=incoming["category_name"]
                                )
                                existing.category = category
                            # Only update fields that were explicitly provided
                            for field in fields_to_compare:
                                if field in provided_fields:
                                    setattr(existing, field, incoming[field])
                            existing.save()
                    else:
                        summary["tests"]["unchanged"] += 1
                        _record_diff(diff, "Tests", str(test_id), "unchanged", {})
                else:
                    summary["tests"]["created"] += 1
                    _record_diff(diff, "Tests", str(test_id), "create", incoming)
                    if not dry_run:
                        category, _ = TestCategory.objects.get_or_create(
                            name=incoming["category_name"]
                        )
                        test = Test.objects.create(
                            test_id=test_id,
                            category=category,
                            **{
                                k: v
                                for k, v in incoming.items()
                                if k != "category_name"
                            },
                        )
                        created_tests[test_id] = test
                    else:
                        # In dry-run, track that this test would be created
                        created_tests[test_id] = True

        # Parameters
        if "Parameters" in workbook.sheetnames:
            sheet = workbook["Parameters"]
            headers = get_headers(sheet)
            seen_ids = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                raw_param_id = safe_get(row, headers, "parameter_id")
                if raw_param_id is None:
                    continue
                try:
                    param_id = _normalize_param_id(raw_param_id)
                except Exception:
                    _add_issue(
                        errors,
                        "Parameters",
                        row_num,
                        "parameter_id",
                        f"Invalid parameter_id {raw_param_id!r}",
                    )
                    continue
                if param_id in seen_ids:
                    _add_issue(
                        errors,
                        "Parameters",
                        row_num,
                        "parameter_id",
                        f"Duplicate parameter_id {param_id}",
                    )
                    continue
                seen_ids.add(param_id)

                param_name = safe_get(row, headers, "parameter_name")
                if not require_value(
                    "Parameters", row_num, "parameter_name", param_name
                ):
                    continue

                # Track which fields were explicitly provided
                provided_fields = {"parameter_name"}
                
                # Get raw values for optional fields
                raw_unit = safe_get(row, headers, "unit")
                raw_data_type = safe_get(row, headers, "data_type")
                raw_editor_type = safe_get(row, headers, "editor_type")
                raw_decimal_places = safe_get(row, headers, "decimal_places")
                raw_allowed_values = safe_get(row, headers, "allowed_values")
                raw_flag_direction = safe_get(row, headers, "flag_direction")
                raw_has_quick_text = safe_get(row, headers, "has_quick_text")
                raw_active = safe_get(row, headers, "active")

                # Process optional fields
                if raw_unit is not None:
                    unit_val = str(raw_unit).strip() or None
                    provided_fields.add("unit")
                else:
                    unit_val = None

                if raw_data_type is not None:
                    data_type_val = str(raw_data_type).strip()
                    provided_fields.add("data_type")
                elif allow_defaults:
                    data_type_val = DEFAULTS["Parameters"]["data_type"]
                    _add_issue(warnings, "Parameters", row_num, "data_type", f"Defaulted to {data_type_val!r}")
                else:
                    data_type_val = None

                if raw_editor_type is not None:
                    editor_type_val = str(raw_editor_type).strip()
                    provided_fields.add("editor_type")
                elif allow_defaults:
                    editor_type_val = DEFAULTS["Parameters"]["editor_type"]
                    _add_issue(warnings, "Parameters", row_num, "editor_type", f"Defaulted to {editor_type_val!r}")
                else:
                    editor_type_val = None

                if raw_decimal_places is not None:
                    decimal_places_val = to_int(raw_decimal_places)
                    if decimal_places_val is not None:
                        provided_fields.add("decimal_places")
                elif allow_defaults:
                    decimal_places_val = DEFAULTS["Parameters"]["decimal_places"]
                    _add_issue(warnings, "Parameters", row_num, "decimal_places", f"Defaulted to {decimal_places_val!r}")
                else:
                    decimal_places_val = None

                if raw_allowed_values is not None:
                    allowed_values_val = str(raw_allowed_values).strip() or ""
                    provided_fields.add("allowed_values")
                else:
                    allowed_values_val = ""

                if raw_flag_direction is not None:
                    flag_direction_val = str(raw_flag_direction).strip()
                    provided_fields.add("flag_direction")
                elif allow_defaults:
                    flag_direction_val = DEFAULTS["Parameters"]["flag_direction"]
                    _add_issue(warnings, "Parameters", row_num, "flag_direction", f"Defaulted to {flag_direction_val!r}")
                else:
                    flag_direction_val = None

                if raw_has_quick_text is not None:
                    has_quick_text_val = to_bool(raw_has_quick_text, DEFAULTS["Parameters"]["has_quick_text"])
                    provided_fields.add("has_quick_text")
                elif allow_defaults:
                    has_quick_text_val = DEFAULTS["Parameters"]["has_quick_text"]
                    _add_issue(warnings, "Parameters", row_num, "has_quick_text", f"Defaulted to {has_quick_text_val!r}")
                else:
                    has_quick_text_val = DEFAULTS["Parameters"]["has_quick_text"]

                if raw_active is not None:
                    active_val = to_bool(raw_active, DEFAULTS["Parameters"]["active"])
                    provided_fields.add("active")
                elif allow_defaults:
                    active_val = DEFAULTS["Parameters"]["active"]
                    _add_issue(warnings, "Parameters", row_num, "active", f"Defaulted to {active_val!r}")
                else:
                    active_val = DEFAULTS["Parameters"]["active"]

                incoming = {
                    "parameter_name": str(param_name).strip(),
                    "unit": unit_val,
                    "data_type": data_type_val,
                    "editor_type": editor_type_val,
                    "decimal_places": decimal_places_val,
                    "allowed_values": allowed_values_val,
                    "flag_direction": flag_direction_val,
                    "has_quick_text": has_quick_text_val if has_quick_text_val is not None else DEFAULTS["Parameters"]["has_quick_text"],
                    "active": active_val if active_val is not None else DEFAULTS["Parameters"]["active"],
                }

                # For new records, ensure required fields have defaults
                if (
                    incoming["data_type"] is None
                    or incoming["editor_type"] is None
                    or incoming["decimal_places"] is None
                ):
                    if strict:
                        if incoming["data_type"] is None:
                            _add_issue(errors, "Parameters", row_num, "data_type", "Missing required value")
                        if incoming["editor_type"] is None:
                            _add_issue(errors, "Parameters", row_num, "editor_type", "Missing required value")
                        if incoming["decimal_places"] is None:
                            _add_issue(errors, "Parameters", row_num, "decimal_places", "Missing required value")
                    continue

                existing = existing_params.get(param_id)
                if existing:
                    # Only compare and update fields that were explicitly provided
                    fields_to_compare = [
                        "parameter_name",
                        "unit",
                        "data_type",
                        "editor_type",
                        "decimal_places",
                        "allowed_values",
                        "flag_direction",
                        "has_quick_text",
                        "active",
                    ]
                    changes = compare_fields(
                        existing,
                        incoming,
                        fields_to_compare,
                        provided_fields=provided_fields,
                    )
                    if changes:
                        summary["parameters"]["updated"] += 1
                        _record_diff(diff, "Parameters", param_id, "update", changes)
                        if not dry_run:
                            # Only update fields that were explicitly provided
                            for field in fields_to_compare:
                                if field in provided_fields:
                                    setattr(existing, field, incoming[field])
                            existing.save()
                    else:
                        summary["parameters"]["unchanged"] += 1
                        _record_diff(diff, "Parameters", param_id, "unchanged", {})
                else:
                    summary["parameters"]["created"] += 1
                    _record_diff(diff, "Parameters", param_id, "create", incoming)
                    if not dry_run:
                        param = Parameter.objects.create(
                            parameter_id=param_id, **incoming
                        )
                        created_params[param_id] = param
                    else:
                        # In dry-run, track that this parameter would be created
                        created_params[param_id] = True

        # Panels
        if "Panels" in workbook.sheetnames:
            sheet = workbook["Panels"]
            headers = get_headers(sheet)
            seen_codes = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                panel_code = safe_get(row, headers, "panel_code")
                if panel_code is None:
                    continue
                panel_code = str(panel_code).strip()
                if panel_code in seen_codes:
                    _add_issue(
                        errors,
                        "Panels",
                        row_num,
                        "panel_code",
                        f"Duplicate panel_code {panel_code}",
                    )
                    continue
                seen_codes.add(panel_code)

                panel_name = safe_get(row, headers, "panel_name")
                category_name = safe_get(row, headers, "category")
                if not require_value("Panels", row_num, "panel_name", panel_name):
                    continue
                if not require_value("Panels", row_num, "category", category_name):
                    continue

                # Track which fields were explicitly provided
                provided_fields = {"panel_code", "panel_name", "category_name"}
                
                # Get raw values for optional fields
                raw_sample_type = safe_get(row, headers, "sample_type")
                raw_price = safe_get(row, headers, "price")
                raw_tat = safe_get(row, headers, "turnaround_time")
                raw_sample_volume = safe_get(row, headers, "sample_volume")
                raw_description = safe_get(row, headers, "description")
                raw_is_active = safe_get(row, headers, "is_active")

                # Process optional fields
                if raw_sample_type is not None:
                    sample_type = str(raw_sample_type).strip()
                    provided_fields.add("sample_type")
                elif allow_defaults:
                    sample_type = DEFAULTS["Panels"]["sample_type"]
                    _add_issue(warnings, "Panels", row_num, "sample_type", f"Defaulted to {sample_type!r}")
                else:
                    sample_type = None

                if raw_price is not None:
                    price_val = to_decimal(raw_price)
                    if price_val is not None:
                        provided_fields.add("price")
                elif allow_defaults:
                    price_val = DEFAULTS["Panels"]["price"]
                    _add_issue(warnings, "Panels", row_num, "price", f"Defaulted to {price_val!r}")
                else:
                    price_val = None

                if raw_tat is not None:
                    tat_val = to_int(raw_tat)
                    if tat_val is not None:
                        provided_fields.add("turnaround_time")
                elif allow_defaults:
                    tat_val = DEFAULTS["Panels"]["turnaround_time"]
                    _add_issue(warnings, "Panels", row_num, "turnaround_time", f"Defaulted to {tat_val!r}")
                else:
                    tat_val = None

                # For required fields with defaults, ensure we have values
                if sample_type is None or price_val is None or tat_val is None:
                    if strict:
                        if sample_type is None:
                            _add_issue(errors, "Panels", row_num, "sample_type", "Missing required value")
                        if price_val is None:
                            _add_issue(errors, "Panels", row_num, "price", "Missing required value")
                        if tat_val is None:
                            _add_issue(errors, "Panels", row_num, "turnaround_time", "Missing required value")
                    continue

                if raw_sample_volume is not None:
                    sample_volume_val = str(raw_sample_volume).strip() or None
                    provided_fields.add("sample_volume")
                else:
                    sample_volume_val = None

                if raw_description is not None:
                    description_val = str(raw_description).strip() or None
                    provided_fields.add("description")
                else:
                    description_val = None

                if raw_is_active is not None:
                    is_active_val = to_bool(raw_is_active, DEFAULTS["Panels"]["is_active"])
                    provided_fields.add("is_active")
                elif allow_defaults:
                    is_active_val = DEFAULTS["Panels"]["is_active"]
                    _add_issue(warnings, "Panels", row_num, "is_active", f"Defaulted to {is_active_val!r}")
                else:
                    is_active_val = DEFAULTS["Panels"]["is_active"]

                incoming = {
                    "panel_code": panel_code,
                    "panel_name": str(panel_name).strip(),
                    "category_name": str(category_name).strip(),
                    "sample_type": str(sample_type).strip(),
                    "sample_volume": sample_volume_val,
                    "price": price_val,
                    "turnaround_time": int(tat_val),
                    "description": description_val,
                    "is_active": is_active_val if is_active_val is not None else DEFAULTS["Panels"]["is_active"],
                }

                existing = existing_panels.get(panel_code)
                if existing:
                    # Only compare and update fields that were explicitly provided
                    fields_to_compare = [
                        "panel_name",
                        "sample_type",
                        "sample_volume",
                        "price",
                        "turnaround_time",
                        "description",
                        "is_active",
                    ]
                    changes = compare_fields(
                        existing,
                        incoming,
                        fields_to_compare,
                        provided_fields=provided_fields,
                    )
                    if "category_name" in provided_fields and existing.category.name != incoming["category_name"]:
                        changes["category"] = {
                            "from": existing.category.name
                            if existing.category
                            else None,
                            "to": incoming["category_name"],
                        }
                    if changes:
                        summary["panels"]["updated"] += 1
                        _record_diff(diff, "Panels", panel_code, "update", changes)
                        if not dry_run:
                            if "category_name" in provided_fields:
                                category, _ = TestCategory.objects.get_or_create(
                                    name=incoming["category_name"]
                                )
                                existing.category = category
                            # Only update fields that were explicitly provided
                            for field in fields_to_compare:
                                if field in provided_fields:
                                    setattr(existing, field, incoming[field])
                            existing.save()
                    else:
                        summary["panels"]["unchanged"] += 1
                        _record_diff(diff, "Panels", panel_code, "unchanged", {})
                else:
                    summary["panels"]["created"] += 1
                    _record_diff(diff, "Panels", panel_code, "create", incoming)
                    if not dry_run:
                        category, _ = TestCategory.objects.get_or_create(
                            name=incoming["category_name"]
                        )
                        panel = TestPanel.objects.create(
                            category=category,
                            **{
                                k: v
                                for k, v in incoming.items()
                                if k not in ["category_name", "panel_code"]
                            },
                            panel_code=panel_code,
                        )
                        created_panels[panel_code] = panel

        # Mapping
        if "Mapping" in workbook.sheetnames:
            sheet = workbook["Mapping"]
            headers = get_headers(sheet)
            seen_pairs = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                test_id = to_int(safe_get(row, headers, "test_id"))
                raw_param_id = safe_get(row, headers, "parameter_id")
                if test_id is None or raw_param_id is None:
                    continue
                try:
                    param_id = _normalize_param_id(raw_param_id)
                except Exception:
                    _add_issue(
                        errors,
                        "Mapping",
                        row_num,
                        "parameter_id",
                        f"Invalid parameter_id {raw_param_id!r}",
                    )
                    continue
                key = (test_id, param_id)
                if key in seen_pairs:
                    _add_issue(
                        errors,
                        "Mapping",
                        row_num,
                        "mapping",
                        f"Duplicate mapping {test_id}:{param_id}",
                    )
                    continue
                seen_pairs.add(key)

                test_obj = created_tests.get(test_id) or existing_tests.get(test_id)
                param_obj = created_params.get(param_id) or existing_params.get(
                    param_id
                )
                if not test_obj or not param_obj:
                    _add_issue(
                        errors,
                        "Mapping",
                        row_num,
                        "mapping",
                        "Test or Parameter not found for mapping",
                    )
                    continue

                # Track which fields were explicitly provided
                provided_fields = set()
                
                # Get raw values for optional fields
                raw_display_order = safe_get(row, headers, "display_order")
                raw_reportable = safe_get(row, headers, "reportable")

                if raw_display_order is not None:
                    display_order_val = to_int(raw_display_order)
                    if display_order_val is not None:
                        provided_fields.add("display_order")
                elif allow_defaults:
                    display_order_val = DEFAULTS["Mapping"]["display_order"]
                    _add_issue(warnings, "Mapping", row_num, "display_order", f"Defaulted to {display_order_val!r}")
                else:
                    display_order_val = None

                if raw_reportable is not None:
                    reportable_val = to_bool(raw_reportable, DEFAULTS["Mapping"]["reportable"])
                    provided_fields.add("reportable")
                elif allow_defaults:
                    reportable_val = DEFAULTS["Mapping"]["reportable"]
                    _add_issue(warnings, "Mapping", row_num, "reportable", f"Defaulted to {reportable_val!r}")
                else:
                    reportable_val = DEFAULTS["Mapping"]["reportable"]

                # For new records, ensure required fields have defaults
                if display_order_val is None:
                    if strict:
                        _add_issue(errors, "Mapping", row_num, "display_order", "Missing required value")
                    continue

                incoming = {
                    "display_order": int(display_order_val),
                    "reportable": reportable_val if reportable_val is not None else True,
                }

                existing = existing_mappings.get((test_id, param_id))
                if existing:
                    # Only compare and update fields that were explicitly provided
                    fields_to_compare = ["display_order", "reportable"]
                    changes = compare_fields(
                        existing, incoming, fields_to_compare, provided_fields=provided_fields
                    )
                    if changes:
                        summary["mappings"]["updated"] += 1
                        _record_diff(
                            diff, "Mapping", f"{test_id}:{param_id}", "update", changes
                        )
                        if not dry_run:
                            # Only update fields that were explicitly provided
                            for field in fields_to_compare:
                                if field in provided_fields:
                                    setattr(existing, field, incoming[field])
                            existing.save()
                    else:
                        summary["mappings"]["unchanged"] += 1
                        _record_diff(
                            diff, "Mapping", f"{test_id}:{param_id}", "unchanged", {}
                        )
                else:
                    summary["mappings"]["created"] += 1
                    _record_diff(
                        diff, "Mapping", f"{test_id}:{param_id}", "create", incoming
                    )
                    if not dry_run:
                        TestParameter.objects.create(
                            test=test_obj,
                            parameter=param_obj,
                            **incoming,
                        )
                    else:
                        created_mappings[(test_id, param_id)] = True

        # PanelTests
        if "PanelTests" in workbook.sheetnames:
            sheet = workbook["PanelTests"]
            headers = get_headers(sheet)
            seen_pairs = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                panel_code = safe_get(row, headers, "panel_code")
                test_id = to_int(safe_get(row, headers, "test_id"))
                if panel_code is None or test_id is None:
                    continue
                panel_code = str(panel_code).strip()
                key = (panel_code, test_id)
                if key in seen_pairs:
                    _add_issue(
                        errors,
                        "PanelTests",
                        row_num,
                        "panel_code",
                        f"Duplicate panel/test {panel_code}:{test_id}",
                    )
                    continue
                seen_pairs.add(key)

                panel_obj = created_panels.get(panel_code) or existing_panels.get(
                    panel_code
                )
                test_obj = created_tests.get(test_id) or existing_tests.get(test_id)
                if not panel_obj or not test_obj:
                    _add_issue(
                        errors,
                        "PanelTests",
                        row_num,
                        "panel_code",
                        "Panel or Test not found for panel mapping",
                    )
                    continue

                existing_rel = panel_obj.tests.filter(test_id=test_id).exists()
                if existing_rel:
                    summary["panel_tests"]["unchanged"] += 1
                    _record_diff(
                        diff, "PanelTests", f"{panel_code}:{test_id}", "unchanged", {}
                    )
                else:
                    summary["panel_tests"]["created"] += 1
                    _record_diff(
                        diff, "PanelTests", f"{panel_code}:{test_id}", "create", {}
                    )
                    if not dry_run:
                        panel_obj.tests.add(test_obj)

        # ReferenceRanges
        if "ReferenceRanges" in workbook.sheetnames:
            sheet = workbook["ReferenceRanges"]
            headers = get_headers(sheet)
            seen_keys = set()
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), 2
            ):
                test_id = to_int(safe_get(row, headers, "test_id"))
                raw_param_id = safe_get(row, headers, "parameter_id")
                if test_id is None or raw_param_id is None:
                    continue
                try:
                    param_id = _normalize_param_id(raw_param_id)
                except Exception:
                    _add_issue(
                        errors,
                        "ReferenceRanges",
                        row_num,
                        "parameter_id",
                        f"Invalid parameter_id {raw_param_id!r}",
                    )
                    continue

                # Track which fields were explicitly provided
                provided_fields = set()
                
                # Get raw values for key fields (gender, age_min, age_max, version)
                raw_gender = safe_get(row, headers, "gender")
                raw_age_min = safe_get(row, headers, "age_min")
                raw_age_max = safe_get(row, headers, "age_max")
                raw_version = safe_get(row, headers, "version")
                
                # Process gender (required for key matching)
                if raw_gender is not None:
                    gender = str(raw_gender).strip()
                    provided_fields.add("gender")
                elif allow_defaults:
                    gender = DEFAULTS["ReferenceRanges"]["gender"]
                    _add_issue(warnings, "ReferenceRanges", row_num, "gender", f"Defaulted to {gender!r}")
                else:
                    gender = DEFAULTS["ReferenceRanges"]["gender"]
                
                if gender not in ["Male", "Female", "Both"]:
                    _add_issue(
                        errors,
                        "ReferenceRanges",
                        row_num,
                        "gender",
                        f"Invalid gender {gender!r}",
                    )
                    continue
                
                # Process age_min and age_max (part of key)
                if raw_age_min is not None:
                    age_min = to_int(raw_age_min)
                    provided_fields.add("age_min")
                else:
                    age_min = None
                
                if raw_age_max is not None:
                    age_max = to_int(raw_age_max)
                    provided_fields.add("age_max")
                else:
                    age_max = None
                
                # Process version (required for key matching)
                if raw_version is not None:
                    version = to_int(raw_version)
                    if version is not None:
                        provided_fields.add("version")
                elif allow_defaults:
                    version = DEFAULTS["ReferenceRanges"]["version"]
                    _add_issue(warnings, "ReferenceRanges", row_num, "version", f"Defaulted to {version}")
                else:
                    version = None
                
                if version is None:
                    if strict:
                        _add_issue(
                            errors,
                            "ReferenceRanges",
                            row_num,
                            "version",
                            "Missing required value (defaults disabled)",
                        )
                    else:
                        _add_issue(
                            warnings,
                            "ReferenceRanges",
                            row_num,
                            "version",
                            "Missing required value (defaults disabled)",
                        )
                    continue

                key = (test_id, param_id, gender, age_min, age_max, version)
                if key in seen_keys:
                    _add_issue(
                        errors,
                        "ReferenceRanges",
                        row_num,
                        "parameter_id",
                        f"Duplicate range {key}",
                    )
                    continue
                seen_keys.add(key)

                test_obj = created_tests.get(test_id) or existing_tests.get(test_id)
                param_obj = created_params.get(param_id) or existing_params.get(
                    param_id
                )
                if not test_obj or not param_obj:
                    _add_issue(
                        errors,
                        "ReferenceRanges",
                        row_num,
                        "parameter_id",
                        "Test or Parameter not found for range",
                    )
                    continue

                mapping = existing_mappings.get(
                    (test_id, param_id)
                ) or created_mappings.get((test_id, param_id))
                if not mapping and not dry_run:
                    mapping = TestParameter.objects.filter(
                        test_id=test_id, parameter_id=param_id
                    ).first()
                if not mapping:
                    _add_issue(
                        errors,
                        "ReferenceRanges",
                        row_num,
                        "parameter_id",
                        "Mapping not found for range",
                    )
                    continue

                # Get raw values for optional fields
                raw_reference_min = safe_get(row, headers, "reference_min")
                raw_reference_max = safe_get(row, headers, "reference_max")
                raw_critical_low = safe_get(row, headers, "critical_low")
                raw_critical_high = safe_get(row, headers, "critical_high")
                raw_is_active = safe_get(row, headers, "is_active")
                raw_notes = safe_get(row, headers, "notes")

                # Process optional fields
                if raw_reference_min is not None:
                    reference_min_val = to_decimal(raw_reference_min)
                    provided_fields.add("reference_min")
                else:
                    reference_min_val = None

                if raw_reference_max is not None:
                    reference_max_val = to_decimal(raw_reference_max)
                    provided_fields.add("reference_max")
                else:
                    reference_max_val = None

                if raw_critical_low is not None:
                    critical_low_val = to_decimal(raw_critical_low)
                    provided_fields.add("critical_low")
                else:
                    critical_low_val = None

                if raw_critical_high is not None:
                    critical_high_val = to_decimal(raw_critical_high)
                    provided_fields.add("critical_high")
                else:
                    critical_high_val = None

                if raw_is_active is not None:
                    is_active_val = to_bool(raw_is_active, DEFAULTS["ReferenceRanges"]["is_active"])
                    provided_fields.add("is_active")
                elif allow_defaults:
                    is_active_val = DEFAULTS["ReferenceRanges"]["is_active"]
                    _add_issue(warnings, "ReferenceRanges", row_num, "is_active", f"Defaulted to {is_active_val!r}")
                else:
                    is_active_val = DEFAULTS["ReferenceRanges"]["is_active"]

                if raw_notes is not None:
                    notes_val = str(raw_notes).strip() or None
                    provided_fields.add("notes")
                else:
                    notes_val = None

                incoming = {
                    "gender": str(gender).strip(),
                    "age_min": age_min,
                    "age_max": age_max,
                    "reference_min": reference_min_val,
                    "reference_max": reference_max_val,
                    "critical_low": critical_low_val,
                    "critical_high": critical_high_val,
                    "is_active": is_active_val if is_active_val is not None else DEFAULTS["ReferenceRanges"]["is_active"],
                    "version": int(version),
                    "notes": notes_val,
                }

                existing = existing_ranges.get(
                    (
                        test_id,
                        param_id,
                        incoming["gender"],
                        age_min,
                        age_max,
                        incoming["version"],
                    )
                )
                if existing:
                    # Only compare and update fields that were explicitly provided
                    # Note: gender, age_min, age_max, version are part of the key, so we don't update them
                    fields_to_compare = [
                        "reference_min",
                        "reference_max",
                        "critical_low",
                        "critical_high",
                        "is_active",
                        "notes",
                    ]
                    changes = compare_fields(
                        existing,
                        incoming,
                        fields_to_compare,
                        provided_fields=provided_fields,
                    )
                    if changes:
                        summary["reference_ranges"]["updated"] += 1
                        _record_diff(
                            diff,
                            "ReferenceRanges",
                            f"{test_id}:{param_id}:{incoming['gender']}:{age_min}:{age_max}:{incoming['version']}",
                            "update",
                            changes,
                        )
                        if not dry_run:
                            # Only update fields that were explicitly provided
                            for field in fields_to_compare:
                                if field in provided_fields:
                                    setattr(existing, field, incoming[field])
                            existing.parameter = mapping
                            existing.save()
                    else:
                        summary["reference_ranges"]["unchanged"] += 1
                        _record_diff(
                            diff,
                            "ReferenceRanges",
                            f"{test_id}:{param_id}:{incoming['gender']}:{age_min}:{age_max}:{incoming['version']}",
                            "unchanged",
                            {},
                        )
                else:
                    summary["reference_ranges"]["created"] += 1
                    _record_diff(
                        diff,
                        "ReferenceRanges",
                        f"{test_id}:{param_id}:{incoming['gender']}:{age_min}:{age_max}:{incoming['version']}",
                        "create",
                        incoming,
                    )
                    if not dry_run:
                        ReferenceRange.objects.create(parameter=mapping, **incoming)

    diff.sort(
        key=lambda d: (
            SHEET_ORDER.index(d["sheet"]) if d["sheet"] in SHEET_ORDER else 99,
            d["key"],
        )
    )

    # Serialize the response to ensure JSON compatibility (convert Decimals to strings)
    return _serialize_for_json(
        {
            "dry_run": dry_run,
            "strict": strict,
            "allow_defaults": allow_defaults,
            "mode": mode,
            "counts": summary,
            "errors": errors,
            "warnings": warnings,
            "diff": diff,
        }
    )


def export_catalog_workbook():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Tests
    tests_sheet = wb.create_sheet("Tests")
    tests_sheet.append(CATALOG_COLUMNS["Tests"])
    for test in Test.objects.select_related("category").order_by("test_id"):
        tests_sheet.append(
            [
                test.test_id,
                test.test_code,
                test.legacy_test_code,
                test.test_name,
                test.category.name if test.category else "",
                test.sample_type,
                test.sample_volume,
                str(test.price),
                test.turnaround_time,
                test.loinc_code,
                test.instructions,
                test.is_active,
            ]
        )

    # Parameters
    params_sheet = wb.create_sheet("Parameters")
    params_sheet.append(CATALOG_COLUMNS["Parameters"])
    for param in Parameter.objects.order_by("parameter_id"):
        params_sheet.append(
            [
                param.parameter_id,
                param.parameter_name,
                param.unit,
                param.data_type,
                param.editor_type,
                param.decimal_places,
                param.allowed_values,
                param.flag_direction,
                param.has_quick_text,
                param.active,
            ]
        )

    # Mapping
    mapping_sheet = wb.create_sheet("Mapping")
    mapping_sheet.append(CATALOG_COLUMNS["Mapping"])
    mappings = TestParameter.objects.select_related("test", "parameter").order_by(
        "test__test_id",
        "display_order",
        "parameter__parameter_id",
    )
    for mapping in mappings:
        mapping_sheet.append(
            [
                mapping.test.test_id,
                mapping.parameter.parameter_id,
                mapping.display_order,
                mapping.reportable,
            ]
        )

    # Panels
    panels_sheet = wb.create_sheet("Panels")
    panels_sheet.append(CATALOG_COLUMNS["Panels"])
    for panel in TestPanel.objects.select_related("category").order_by("panel_code"):
        panels_sheet.append(
            [
                panel.panel_code,
                panel.panel_name,
                panel.category.name if panel.category else "",
                panel.sample_type,
                panel.sample_volume,
                str(panel.price),
                panel.turnaround_time,
                panel.description,
                panel.is_active,
            ]
        )

    # PanelTests
    panel_tests_sheet = wb.create_sheet("PanelTests")
    panel_tests_sheet.append(CATALOG_COLUMNS["PanelTests"])
    panels = TestPanel.objects.prefetch_related("tests").order_by("panel_code")
    for panel in panels:
        tests = sorted(panel.tests.all(), key=lambda t: t.test_id)
        for test in tests:
            panel_tests_sheet.append(
                [
                    panel.panel_code,
                    test.test_id,
                ]
            )

    # ReferenceRanges
    ranges_sheet = wb.create_sheet("ReferenceRanges")
    ranges_sheet.append(CATALOG_COLUMNS["ReferenceRanges"])
    ranges = ReferenceRange.objects.select_related(
        "parameter__test",
        "parameter__parameter",
    ).order_by(
        "parameter__test__test_id",
        "parameter__parameter__parameter_id",
        "gender",
        "age_min",
        "age_max",
        "version",
    )
    for rr in ranges:
        ranges_sheet.append(
            [
                rr.parameter.test.test_id,
                rr.parameter.parameter.parameter_id,
                rr.gender,
                rr.age_min,
                rr.age_max,
                str(rr.reference_min) if rr.reference_min is not None else None,
                str(rr.reference_max) if rr.reference_max is not None else None,
                str(rr.critical_low) if rr.critical_low is not None else None,
                str(rr.critical_high) if rr.critical_high is not None else None,
                rr.is_active,
                rr.version,
                rr.notes,
            ]
        )

    return wb


class DummyContext:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False
