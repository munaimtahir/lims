import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tests"

ws.append(["test_id", "test_code", "test_name", "category", "sample_type", "price", "turnaround_time", "is_active"])
ws.append([1000, "TEST1", "Test 1", "Category 1", "Serum", 100, 24, True])

ws = wb.create_sheet("Parameters")
ws.append(["parameter_id", "parameter_name", "unit", "data_type", "editor_type", "decimal_places", "allowed_values", "flag_direction", "has_quick_text", "active"])
ws.append(["p1000", "Param 1", "mg/dL", "Numeric", "Plain", 2, "", "Both", False, True])

ws = wb.create_sheet("Mapping")
ws.append(["test_id", "parameter_id", "display_order", "reportable"])
ws.append([1000, "p1000", 1, True])

wb.save("test_catalog.xlsx")
