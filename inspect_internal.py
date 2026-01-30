import openpyxl
import sys

def inspect(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"\nSheet: {sheet}")
            rows = list(ws.iter_rows(max_row=3, values_only=True))
            for i, r in enumerate(rows):
                print(f"Row {i+1}: {r}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        inspect(sys.argv[1])
    else:
        print("Usage: inspect.py <file>")
