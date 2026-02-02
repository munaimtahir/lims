import openpyxl
import sys

file_path = sys.argv[1]

wb = openpyxl.load_workbook(file_path)

ws = wb["Parameters"]

headers = []
for cell in ws[1]:
    headers.append(cell.value)

if "data_type" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "data_type"
    headers.append("data_type")
if "editor_type" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "editor_type"
    headers.append("editor_type")
if "decimal_places" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "decimal_places"
    headers.append("decimal_places")
if "flag_direction" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "flag_direction"
    headers.append("flag_direction")
if "has_quick_text" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "has_quick_text"
    headers.append("has_quick_text")
if "active" not in headers:
    ws.cell(row=1, column=len(headers) + 1).value = "active"
    headers.append("active")


for row in ws.iter_rows(min_row=2):
    row[headers.index("data_type")].value = "Numeric"
    row[headers.index("editor_type")].value = "Plain"
    row[headers.index("decimal_places")].value = 2
    row[headers.index("flag_direction")].value = "Both"
    row[headers.index("has_quick_text")].value = False
    row[headers.index("active")].value = True

wb.save(file_path)
