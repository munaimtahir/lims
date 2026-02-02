import openpyxl
import sys

try:
    file_path = sys.argv[1]
    wb = openpyxl.load_workbook(file_path)
    
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # Get the first row (headers)
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        print(f"Columns: {headers}")
        
except Exception as e:
    print(f"Error reading excel: {e}")
